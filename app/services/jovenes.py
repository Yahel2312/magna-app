import os
from sqlalchemy.orm import Session
from openpyxl import load_workbook

import app.models as models

GRUPOS = ["Secundarios", "Prepos", "Universitarios", "Profesionistas"]


def buscar_jovenes(nombre: str, db: Session) -> list[models.Joven]:
    """Búsqueda case-insensitive de jóvenes por nombre."""
    return (
        db.query(models.Joven)
        .filter(models.Joven.nombre.ilike(f"%{nombre}%"))
        .order_by(models.Joven.nombre)
        .all()
    )


def obtener_estadisticas_por_grupo(evento_id: int, db: Session) -> dict[str, int]:
    """
    Retorna el conteo de asistentes por grupo para un evento dado.
    Usa un JOIN en lugar de consultas N+1.
    """
    grupos: dict[str, int] = {g: 0 for g in GRUPOS}
    grupos["Sin grupo"] = 0

    rows = (
        db.query(models.Joven.grupo)
        .join(models.Asistencia, models.Asistencia.joven_id == models.Joven.id)
        .filter(models.Asistencia.evento_id == evento_id)
        .all()
    )

    for (grupo,) in rows:
        key = grupo or "Sin grupo"
        if key not in grupos:
            grupos[key] = 0
        grupos[key] += 1

    return grupos


def obtener_top_rachas(db: Session, limit: int = 10) -> list[models.Joven]:
    """Retorna los jóvenes con la racha actual más alta."""
    return (
        db.query(models.Joven)
        .filter(models.Joven.racha_actual > 0)
        .order_by(models.Joven.racha_actual.desc(), models.Joven.puntos_totales.desc())
        .limit(limit)
        .all()
    )


def obtener_historial(db: Session, semanas: int = 8) -> list[dict]:
    """
    Retorna el historial de asistencia de los últimos N eventos (semanas),
    con conteo total y por grupo para cada uno.
    """
    eventos = (
        db.query(models.Evento)
        .order_by(models.Evento.fecha.desc())
        .limit(semanas)
        .all()
    )

    historial = []
    for evento in eventos:
        total = (
            db.query(models.Asistencia)
            .filter(models.Asistencia.evento_id == evento.id)
            .count()
        )
        por_grupo = obtener_estadisticas_por_grupo(evento.id, db)
        historial.append(
            {
                "evento_id": evento.id,
                "fecha": evento.fecha.strftime("%d/%m/%Y"),
                "total": total,
                "por_grupo": por_grupo,
            }
        )

    return historial


def importar_desde_excel(db: Session, base_dir: str) -> None:
    """
    Lee Chicos.xlsx y crea/actualiza jóvenes en la base de datos.
    Columnas: A=Secundarios, B=Prepos, C=Universitarios, D=Profesionistas.
    Es idempotente: no duplica jóvenes existentes.
    """
    archivo = os.path.join(base_dir, "Chicos.xlsx")
    if not os.path.exists(archivo):
        return

    wb = load_workbook(archivo)
    ws = wb.active

    for fila in ws.iter_rows(min_row=2, values_only=True):
        for i, nombre in enumerate(fila[:4]):
            if not nombre or not str(nombre).strip():
                continue

            nombre_limpio = str(nombre).strip()
            grupo = GRUPOS[i]

            existente = (
                db.query(models.Joven)
                .filter(models.Joven.nombre.ilike(nombre_limpio))
                .first()
            )

            if existente:
                existente.grupo = grupo
            else:
                db.add(
                    models.Joven(
                        nombre=nombre_limpio,
                        grupo=grupo,
                        puntos_totales=0,
                        puntos_racha=0,
                        racha_actual=0,
                        racha_maxima=0,
                    )
                )

    db.commit()
