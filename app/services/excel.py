import os
from datetime import datetime

from fastapi import HTTPException
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

import app.models as models

GRUPOS = ["Secundarios", "Prepos", "Universitarios", "Profesionistas"]

COLORES_GRUPO = {
    "Secundarios": "BBDEFB",
    "Prepos": "E1BEE7",
    "Universitarios": "B2EBF2",
    "Profesionistas": "FFE0B2",
    "Sin grupo": "F5F5F5",
}


def _aplicar_encabezado(ws, fila: int, columnas: list[str]) -> None:
    for col, titulo in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=col, value=titulo)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1A3A40")
        celda.alignment = Alignment(horizontal="center", vertical="center")


def generar_excel_evento(evento_id: int, db: Session, base_dir: str) -> FileResponse:
    """
    Genera un Excel del evento seleccionado.
    """

    evento = (
        db.query(models.Evento)
        .filter(models.Evento.id == evento_id)
        .first()
    )

    if not evento:
        raise HTTPException(status_code=404, detail="Evento no encontrado")

    filas = (
        db.query(models.Asistencia, models.Joven)
        .join(
            models.Joven,
            models.Asistencia.joven_id == models.Joven.id
        )
        .filter(models.Asistencia.evento_id == evento_id)
        .order_by(models.Joven.grupo, models.Joven.nombre)
        .all()
    )

    # ==========================================================
    # DEBUG
    # ==========================================================
    print("\n" + "=" * 70)
    print("GENERANDO EXCEL")
    print("Evento solicitado:", evento_id)
    print("Fecha del evento:", evento.fecha)
    print("Número de asistencias encontradas:", len(filas))

    for asistencia, joven in filas:
        print(
            f"Asistencia ID={asistencia.id} | "
            f"Evento={asistencia.evento_id} | "
            f"Joven={joven.id} - {joven.nombre} | "
            f"Hora={asistencia.fecha_hora}"
        )

    print("=" * 70 + "\n")
    # ==========================================================

    conteo = {g: 0 for g in GRUPOS}
    conteo["Sin grupo"] = 0

    registros = []

    for asistencia, joven in filas:
        grupo = joven.grupo or "Sin grupo"

        if grupo not in conteo:
            conteo[grupo] = 0

        conteo[grupo] += 1

        registros.append(
            (
                joven.nombre,
                grupo,
                asistencia.fecha_hora,
            )
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Magna App — Registro de Asistencia"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0B2F35")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Fecha: {evento.fecha.strftime('%d/%m/%Y')}"
    ws["A3"] = f"Total asistentes: {len(registros)}"

    ws["A2"].font = Font(bold=True)
    ws["A3"].font = Font(bold=True)

    ws["A5"] = "Resumen por grupo"
    ws["A5"].font = Font(bold=True, size=11)

    fila = 6

    for grupo, total in conteo.items():
        ws.cell(row=fila, column=1, value=grupo)
        ws.cell(row=fila, column=2, value=total)

        color = COLORES_GRUPO.get(grupo, "FFFFFF")

        ws.cell(row=fila, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=fila, column=2).fill = PatternFill("solid", fgColor=color)

        fila += 1

    fila_tabla = fila + 1

    _aplicar_encabezado(
        ws,
        fila_tabla,
        ["#", "Nombre", "Grupo", "Hora de registro"],
    )

    for i, (nombre, grupo, fecha_hora) in enumerate(registros, start=1):
        fila_tabla += 1

        ws.cell(row=fila_tabla, column=1, value=i)
        ws.cell(row=fila_tabla, column=2, value=nombre)
        ws.cell(row=fila_tabla, column=3, value=grupo)
        ws.cell(
            row=fila_tabla,
            column=4,
            value=fecha_hora.strftime("%H:%M:%S") if fecha_hora else "",
        )

        color = COLORES_GRUPO.get(grupo, "FFFFFF")

        for col in range(1, 5):
            ws.cell(row=fila_tabla, column=col).fill = PatternFill(
                "solid",
                fgColor=color,
            )
            ws.cell(row=fila_tabla, column=col).alignment = Alignment(
                horizontal="center"
            )

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 22

    nombre_archivo = (
        f"asistencia_evento_{evento_id}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    ruta = os.path.join(base_dir, nombre_archivo)

    wb.save(ruta)

    return FileResponse(
        ruta,
        filename=nombre_archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )