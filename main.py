from fastapi import FastAPI, Depends
from sqlalchemy.orm import Session
from database import SessionLocal, engine
from pydantic import BaseModel
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from fastapi.responses import FileResponse
from sqlalchemy import text
import os
import models

# Crear las tablas
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Asistencia Juvenil Gamificada")
# ---------------------------
# MODELO PARA ENTRADA DE DATOS
# ---------------------------
class JovenCreate(BaseModel):
    nombre: str

# ---------------------------
# DEPENDENCIA PARA LA DB
# ---------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def importar_chicos_automatico(db):

    archivo = os.path.join(BASE_DIR, "Chicos.xlsx")

    wb = load_workbook(archivo)
    ws = wb.active

    grupos = [
        "Secundarios",
        "Prepos",
        "Universitarios",
        "Profesionistas"
    ]

    for fila in ws.iter_rows(min_row=2, values_only=True):

        for i, nombre in enumerate(fila[:4]):

            if nombre and str(nombre).strip():

                nombre_limpio = str(nombre).strip()
                grupo = grupos[i]

                existe = db.query(models.Joven).filter(
                    models.Joven.nombre.ilike(nombre_limpio)
                ).first()

                if not existe:

                    nuevo = models.Joven(
                        nombre=nombre_limpio,
                        grupo=grupo,
                        puntos_totales=0,
                        puntos_racha=0,
                        racha_actual=0,
                        racha_maxima=0
                    )

                    db.add(nuevo)

                else:
                    existe.grupo = grupo

    db.commit()
def migrar_grupo_si_no_existe(db):
    db.execute(text("""
        ALTER TABLE jovenes
        ADD COLUMN IF NOT EXISTS grupo VARCHAR DEFAULT 'Sin grupo'
    """))
    db.commit()

@app.on_event("startup")
def startup_event():
    db = SessionLocal()

    try:
        migrar_grupo_si_no_existe(db)
        importar_chicos_automatico(db)
    finally:
        db.close()
def obtener_o_crear_evento(db):
    hoy = datetime.now().date()

    evento = db.query(models.Evento).filter(
        models.Evento.fecha >= datetime.combine(hoy, datetime.min.time()),
        models.Evento.fecha <= datetime.combine(hoy, datetime.max.time())
    ).first()

    if evento:
        return evento

    nuevo_evento = models.Evento(
        fecha=datetime.now(),
        activo=True
    )

    db.add(nuevo_evento)
    db.commit()
    db.refresh(nuevo_evento)

    return nuevo_evento

@app.get("/conteo")
def conteo(db: Session = Depends(get_db)):
    total = db.query(models.Joven).count()
    return {"total_jovenes": total}
# ---------------------------
# ENDPOINT REGISTRAR JOVEN
# ---------------------------
@app.post("/jovenes")
def crear_joven(joven: JovenCreate, db: Session = Depends(get_db)):
    nuevo_joven = models.Joven(
    nombre=joven.nombre,
    puntos_totales=0,
    puntos_racha=0,
    racha_actual=0,
    racha_maxima=0
)
    db.add(nuevo_joven)
    db.commit()
    db.refresh(nuevo_joven)
    return {
        "mensaje": "Joven registrado",
        "id": nuevo_joven.id,
        "nombre": nuevo_joven.nombre
    }

@app.post("/eventos")
def crear_evento(db: Session = Depends(get_db)):
    nuevo_evento = models.Evento(
        fecha=datetime.now(),
        activo=True
    )
    db.add(nuevo_evento)
    db.commit()
    db.refresh(nuevo_evento)
    return {
        "mensaje": "Evento creado",
        "evento_id": nuevo_evento.id
    }


@app.post("/asistencia")
def registrar_asistencia(joven_id: int, evento_id: int, db: Session = Depends(get_db)):

    joven = db.query(models.Joven).filter(models.Joven.id == joven_id).first()
    if not joven:
        return {"error": "Joven no encontrado"}

    evento = db.query(models.Evento).filter(models.Evento.id == evento_id).first()
    if not evento:
        return {"error": "Evento no encontrado"}

    # evitar doble registro
    existe = db.query(models.Asistencia).filter(
        models.Asistencia.joven_id == joven_id,
        models.Asistencia.evento_id == evento_id
    ).first()

    if existe:
        return {"mensaje": "Asistencia ya registrada"}

    # buscar evento anterior
    evento_anterior = db.query(models.Evento)\
        .filter(models.Evento.id < evento_id)\
        .order_by(models.Evento.id.desc())\
        .first()

    # verificar si asistió al evento anterior
    asistio_anterior = False

    if evento_anterior:
        asistencia_anterior = db.query(models.Asistencia).filter(
            models.Asistencia.joven_id == joven_id,
            models.Asistencia.evento_id == evento_anterior.id
        ).first()

        if asistencia_anterior:
            asistio_anterior = True

    # registrar asistencia
    nueva = models.Asistencia(
        joven_id=joven_id,
        evento_id=evento_id
    )
    db.add(nueva)

    # ---- GAMIFICACIÓN ----
    if asistio_anterior:
       joven.racha_actual += 1
       joven.puntos_racha += 10
    else:
       joven.racha_actual = 1
       joven.puntos_racha = 10

# SIEMPRE suma puntos totales
    joven.puntos_totales += 10

    db.commit()

    return {
        "mensaje": "Asistencia registrada",
        "puntos": joven.puntos_totales,
        "racha": joven.racha_actual
    }
    
    
@app.get("/joven/{joven_id}")
def ver_joven(joven_id: int, db: Session = Depends(get_db)):
        joven = db.query(models.Joven).filter(models.Joven.id == joven_id).first()
    
        if not joven:
            return {"error": "No existe"}
    
        return {
            "nombre": joven.nombre,
            "puntos_totales": joven.puntos_totales,
            "puntos_racha": joven.puntos_racha,
            "racha_actual": joven.racha_actual,
            "racha_maxima": joven.racha_maxima
        }

@app.get("/admin/jovenes")
def ver_todos(db: Session = Depends(get_db)):
    jovenes = db.query(models.Joven).all()

    resultado = []

    for j in jovenes:
        resultado.append({
            "id": j.id,
            "nombre": j.nombre,
            "puntos_totales": j.puntos_totales,
            "puntos_racha": j.puntos_racha,
            "racha_actual": j.racha_actual,
            "racha_maxima": j.racha_maxima
        })

    return resultado
@app.get("/buscar")
def buscar(nombre: str, db: Session = Depends(get_db)):
    resultados = db.query(models.Joven).filter(
        models.Joven.nombre.ilike(f"%{nombre}%")
    ).all()

    return [
        {"id": j.id, "nombre": j.nombre}
        for j in resultados
    ]
def generar_excel(db):
    asistencias = db.query(models.Asistencia).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"

    ws.append(["Nombre", "Evento ID", "Fecha"])

    for a in asistencias:
        joven = db.query(models.Joven).filter(
            models.Joven.id == a.joven_id
        ).first()

@app.post("/asistencia_manual")
def asistencia_manual(joven_id: int, evento_id: int, db: Session = Depends(get_db)):

    joven = db.query(models.Joven).filter(models.Joven.id == joven_id).first()
    if not joven:
        return {"error": "No encontrado"}

    evento = db.query(models.Evento).filter(models.Evento.id == evento_id).first()
    if not evento:
        return {"error": "Evento no encontrado"}

    existe = db.query(models.Asistencia).filter(
        models.Asistencia.joven_id == joven_id,
        models.Asistencia.evento_id == evento_id
    ).first()

    if not existe:
        nueva = models.Asistencia(joven_id=joven_id, evento_id=evento_id)
        db.add(nueva)

        joven.puntos_totales += 10
        db.commit()

    generar_excel(db)

    return {"mensaje": "Asistencia registrada"}



@app.get("/evento/{evento_id}/conteo")
def conteo_evento(evento_id: int, db: Session = Depends(get_db)):
    total = db.query(models.Asistencia).filter(
        models.Asistencia.evento_id == evento_id
    ).count()

    return {"asistentes": total}

@app.get("/evento/activo")
def evento_activo(db: Session = Depends(get_db)):
    evento = obtener_o_crear_evento(db)

    return {
        "evento_id": evento.id,
        "fecha": str(evento.fecha)
    }
    evento = db.query(models.Evento)\
        .order_by(models.Evento.id.desc())\
        .first()

    if not evento:
        return {"error": "No hay eventos"}

    return {"evento_id": evento.id}

def exportar(evento_id: int, db: Session = Depends(get_db)):

    asistencias = db.query(models.Asistencia).filter(
        models.Asistencia.evento_id == evento_id
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    # encabezados
    ws.append(["Nombre", "Fecha"])

    for a in asistencias:
        joven = db.query(models.Joven).filter(
            models.Joven.id == a.joven_id
        ).first()

        ws.append([
            joven.nombre,
            str(a.fecha_hora)
        ])

    archivo = "asistencia.xlsx"
    wb.save(archivo)

    return FileResponse(archivo, filename=archivo)

@app.get("/")
def home():
    return FileResponse(os.path.join(BASE_DIR, "main.html"))


@app.get("/admin/excel")
def ver_excel(db: Session = Depends(get_db)):

    asistencias = db.query(models.Asistencia).all()

    wb = Workbook()
    ws = wb.active

    ws.title = "Asistencias"

    ws.append([
        "Nombre",
        "Evento",
        "Fecha"
    ])

    for a in asistencias:

        joven = db.query(models.Joven).filter_by(id=a.joven_id).first()

        evento = db.query(models.Evento).filter_by(id=a.evento_id).first()

        ws.append([
            joven.nombre,
            evento.nombre if evento else "Sin evento",
            str(a.fecha_hora)
        ])

    nombre = f"asistencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    ruta = os.path.join(BASE_DIR, nombre)

    wb.save(ruta)

    return FileResponse(
        ruta,
        filename=nombre

    )

@app.get("/admin/excel/asistentes")
def excel_asistentes(db: Session = Depends(get_db)):

    asistencias = db.query(models.Asistencia).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistentes"

    ws.append(["Nombre", "Evento ID", "Fecha"])

    usados = set()

    for a in asistencias:
        clave = (a.joven_id, a.evento_id)

        if clave in usados:
            continue

        usados.add(clave)

        joven = db.query(models.Joven).filter(
            models.Joven.id == a.joven_id
        ).first()

        if joven:
            ws.append([
                joven.nombre,
                a.evento_id,
                str(a.fecha_hora)
            ])

    nombre = f"asistentes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = os.path.join(BASE_DIR, nombre)

    wb.save(ruta)

    return FileResponse(ruta, filename=nombre)


@app.get("/debug/asistencias")
def debug_asistencias(db: Session = Depends(get_db)):
    asistencias = db.query(models.Asistencia).all()

    resultado = []

    for a in asistencias:
        joven = db.query(models.Joven).filter(
            models.Joven.id == a.joven_id
        ).first()

        resultado.append({
            "asistencia_id": a.id,
            "joven_id": a.joven_id,
            "nombre": joven.nombre if joven else "NO ENCONTRADO",
            "evento_id": a.evento_id,
            "fecha": str(a.fecha_hora)
        })

    return {
        "total_asistencias": len(resultado),
        "registros": resultado
    }
@app.get("/admin/excel/evento/{evento_id}")
def excel_evento(evento_id: int, db: Session = Depends(get_db)):

    evento = db.query(models.Evento).filter(
        models.Evento.id == evento_id
    ).first()

    if not evento:
        return {"error": "Evento no encontrado"}

    asistencias = db.query(models.Asistencia).filter(
        models.Asistencia.evento_id == evento_id
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    ws["A1"] = "Registro de Asistencia"
    ws["A2"] = f"Evento ID: {evento.id}"
    ws["A3"] = f"Fecha del evento: {evento.fecha}"
    ws["A4"] = f"Total asistentes: {len(asistencias)}"

    ws.append([])
    ws.append(["No.", "Nombre", "Hora de registro"])

    fila_num = 1

    for a in asistencias:
        joven = db.query(models.Joven).filter(
            models.Joven.id == a.joven_id
        ).first()

        if joven:
            ws.append([
                fila_num,
                joven.nombre,
                str(a.fecha_hora)
            ])
            fila_num += 1

    nombre = f"asistencia_evento_{evento_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    ruta = os.path.join(BASE_DIR, nombre)

    wb.save(ruta)

    return FileResponse(ruta, filename=nombre)
@app.get("/admin/excel/activo")
def excel_evento_activo(db: Session = Depends(get_db)):

    evento = obtener_o_crear_evento(db)

    return excel_evento(evento.id, db)

@app.get("/admin/migrar-grupo")
def migrar_grupo(db: Session = Depends(get_db)):
    db.execute(text("ALTER TABLE jovenes ADD COLUMN IF NOT EXISTS grupo VARCHAR DEFAULT 'Sin grupo'"))
    db.commit()
    return {"mensaje": "Columna grupo agregada"}

@app.get("/admin/migrar-grupo")
def migrar_grupo(db: Session = Depends(get_db)):
    db.execute(text("""
        ALTER TABLE jovenes
        ADD COLUMN IF NOT EXISTS grupo VARCHAR DEFAULT 'Sin grupo'
    """))
    db.commit()

    return {"mensaje": "Columna grupo creada correctamente"}

